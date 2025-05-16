#!/usr/bin/env python
# coding: utf-8

# In[1]:


import arcpy
import os
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# In[2]:


# Shapefiles
folder_path = r"P:\PJ00356 - Inch Cape\356a Survey\87 GIS\7_Deliverable Prep\ECR REV1.0\Greenlight deliverables\SSS\SBF"
# CSVs
csv_output_folder = r"D:\QC_09-05\attributetable_output.xlsx"
# Output
final_output_folder = r"D:\QC_09-05\attributetable_output_themissingdatacolumns.xlsx"


# In[12]:


# Create the output folders if they don't exist
if not os.path.exists(csv_output_folder):
   os.makedirs(csv_output_folder)
if not os.path.exists(final_output_folder):
   os.makedirs(final_output_folder)

# Set the environment workspace 
arcpy.env.workspace = folder_path

# List all shapefiles in the folder
shapefiles = arcpy.ListFeatureClasses()


for shapefile in shapefiles:
   try:
       
        # Get the shapefile's full path
       shapefile_path = os.path.join(folder_path, shapefile)
       
        # Define the output CSV file path
       csv_output_file = os.path.join(csv_output_folder, f"{os.path.splitext(shapefile)[0]}_attribute_table.csv")
       
        # Open the CSV file for writing
       with open(csv_output_file, 'w', newline='') as csvfile:
           writer = csv.writer(csvfile)
           
        #  get the fields and rows from the shapefile
           fields = [field.name for field in arcpy.ListFields(shapefile)]
           writer.writerow(fields)  # Write the field names as the first row
           
            #  access the attribute data
           with arcpy.da.SearchCursor(shapefile_path, fields) as cursor:
               for row in cursor:
                   writer.writerow(row)  # Write each row to the CSV file
       print(f"Attribute table for {shapefile} exported to {csv_output_file}")
       
        # Convert the CSV to Excel 
       excel_output_file = csv_output_file.replace('.csv', '.xlsx')
       df = pd.read_csv(csv_output_file)
       df.to_excel(excel_output_file, index=False)
      
        # Load the Excel file 
       wb = load_workbook(excel_output_file)
       ws = wb.active
      
    # redfill
       red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
       
    # FInd blank cells
       for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
           for cell in row:
               if cell.value is None or str(cell.value).strip() == "":
                   cell.fill = red_fill
       
    # Define the final output path QC spreadsheet
       final_output_file = os.path.join(final_output_folder, f"{os.path.splitext(shapefile)[0]}_highlighted.xlsx")
       
    # Save the workbook with the highlighted cells
       wb.save(final_output_file)
       print(f"Data extracted and saved to {final_output_file} with highlighted empty cells.")
   except Exception as e:
       print(f"Error processing {shapefile}: {e}")


# In[ ]:




