import os
import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Function to extract metadata from XML files
def extract_data_from_xml(xml_file):
    try:
        # Parse the XML file
        tree = etree.parse(xml_file)
        root = tree.getroot()

        # Define namespaces
        ns = {
            'gmd': 'http://www.isotc211.org/2005/gmd',
            'gco': 'http://www.isotc211.org/2005/gco',
            'cit': 'http://www.isotc211.org/2005/cit'
        }

        # Extract basic metadata
        data = {
            'Date': root.findtext('.//gmd:dateStamp/gco:Date', namespaces=ns),
            'File_name': root.findtext('.//gmd:title/gco:CharacterString', namespaces=ns),
            'Abstract': root.findtext('.//gmd:abstract/gco:CharacterString', namespaces=ns),
            'useLimitation': root.findtext('.//gmd:useLimitation/gco:CharacterString', namespaces=ns),
            'Lineage': root.findtext('.//gmd:statement/gco:CharacterString', namespaces=ns),
            'West_bound_longitude': root.findtext('.//gmd:westBoundLongitude/gco:Decimal', namespaces=ns),
            'East_bound_longitude': root.findtext('.//gmd:eastBoundLongitude/gco:Decimal', namespaces=ns),
            'North_bound_latitude': root.findtext('.//gmd:northBoundLatitude/gco:Decimal', namespaces=ns),
            'South_bound_latitude': root.findtext('.//gmd:southBoundLatitude/gco:Decimal', namespaces=ns),
            'General Contact_email': root.findtext('.//gmd:electronicMailAddress/gco:CharacterString', namespaces=ns)
        }

        # Extract roles for originator, owner, and custodian
        role_codes = {
            'Contact of the Originator': None,
            'Contact of the Owner': None,
            'Contact of the Custodian': None
        }

        for role in root.xpath('.//gmd:CI_ResponsibleParty', namespaces=ns):
            role_code = role.findtext('.//gmd:CI_RoleCode', namespaces=ns)
            role_description = role.findtext('.//gmd:electronicMailAddress/gco:CharacterString', namespaces=ns)

            if role_code == 'originator':
                role_codes['Contact of the Originator'] = role_description
            elif role_code == 'owner':
                role_codes['Contact of the Owner'] = role_description
            elif role_code == 'custodian':
                role_codes['Contact of the Custodian'] = role_description

        data.update(role_codes)
        return data

    except Exception as e:
        print(f"Error processing file {xml_file}: {str(e)}")
        return None


# Define folder path and get XML files
folder_path = r"P:\PJ00356 - Inch Cape\356a Survey\87 GIS\3_Metadata\MEDIN\final\MBES"
xml_files = [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.endswith(".xml")]

# Extract metadata from all XML files
all_data = []

for xml_file in xml_files:
    metadata = extract_data_from_xml(xml_file)
    if metadata:  # Only add if data extraction was successful
        all_data.append(metadata)

# Create DataFrame from extracted metadata
if all_data:
    df = pd.DataFrame(all_data)

    # Function to check row completeness
    def check_row_status(row):
        if all(pd.isnull(value) or str(value).strip() == "" for value in row):
            return None
        for value in row:
            if pd.isnull(value) or str(value).strip() == "":
                return "Fail"
        return "Pass"

    # Apply check and remove empty rows
    df['Status'] = df.apply(check_row_status, axis=1)
    df.dropna(subset=['Status'], inplace=True)

    # Save DataFrame to Excel file
    output_file = r"P:\PJ00356 - Inch Cape\356a Survey\87 GIS\7_Deliverable Prep\QC\MBES_1_metadata.xlsx"
    df.to_excel(output_file, index=False)

    # Load workbook and apply red highlights for missing values
    wb = load_workbook(output_file)
    ws = wb.active

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column - 1):
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                cell.fill = red_fill

    # Save the workbook
    wb.save(output_file)
    print(f'Data extracted and saved to {output_file} with highlighted empty cells and pass/fail status column.')
else:
    print("No valid metadata extracted from XML files.")
